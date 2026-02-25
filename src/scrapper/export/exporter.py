"""Export leads to CSV, JSON, or Excel."""

from __future__ import annotations

import csv
import json
from datetime import datetime
from typing import TYPE_CHECKING, Any

if TYPE_CHECKING:
    from pathlib import Path

from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from scrapper.config import get_settings
from scrapper.db.connection import get_connection

# Column definitions for Excel
LEAD_COLUMNS: list[tuple[str, str, int]] = [
    ("company_name", "Компания", 40),
    ("inn", "ИНН", 15),
    ("ogrn", "ОГРН", 18),
    ("kpp", "КПП", 12),
    ("address", "Адрес", 50),
    ("ceo", "Руководитель", 30),
    ("phone", "Телефон", 20),
    ("email", "Email", 30),
    ("website", "Сайт", 25),
    ("revenue", "Выручка", 20),
    ("employees", "Сотрудники", 15),
    ("status", "Статус", 15),
    ("okved", "ОКВЭД", 30),
    ("registration_date", "Дата рег.", 15),
]

# Styles
_HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _fetch_leads(limit: int | None = None) -> list[dict[str, Any]]:
    """Fetch leads from database."""
    query = "SELECT * FROM leads ORDER BY created_at DESC"
    if limit:
        query += f" LIMIT {limit}"
    with get_connection() as conn:
        rows = conn.execute(query).fetchall()
    return [dict(r) for r in rows]


def export_csv(output: Path | None = None, limit: int | None = None) -> Path:
    """Export leads to CSV."""
    leads = _fetch_leads(limit)
    settings = get_settings()
    if output is None:
        output = settings.export_dir / f"leads_{datetime.now():%Y%m%d_%H%M%S}.csv"
    output.parent.mkdir(parents=True, exist_ok=True)

    fieldnames = [col[0] for col in LEAD_COLUMNS]
    with open(output, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for lead in leads:
            writer.writerow(lead)

    logger.info(f"Exported {len(leads)} leads to CSV: {output}")
    return output


def export_json(output: Path | None = None, limit: int | None = None, include_raw: bool = False) -> Path:
    """Export leads to JSON."""
    leads = _fetch_leads(limit)
    settings = get_settings()
    if output is None:
        output = settings.export_dir / f"leads_{datetime.now():%Y%m%d_%H%M%S}.json"
    output.parent.mkdir(parents=True, exist_ok=True)

    if not include_raw:
        for lead in leads:
            lead.pop("raw_data", None)

    # Convert datetimes to strings
    def _serialize(obj: Any) -> Any:
        if isinstance(obj, datetime):
            return obj.isoformat()
        return str(obj)

    with open(output, "w", encoding="utf-8") as f:
        json.dump(leads, f, ensure_ascii=False, indent=2, default=_serialize)

    logger.info(f"Exported {len(leads)} leads to JSON: {output}")
    return output


def export_xlsx(output: Path | None = None, limit: int | None = None) -> Path:
    """Export leads to Excel (.xlsx) with formatting."""
    leads = _fetch_leads(limit)
    settings = get_settings()
    if output is None:
        output = settings.export_dir / f"leads_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    output.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    # Write headers
    for col_idx, (_field, header, width) in enumerate(LEAD_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Write data
    for row_idx, lead in enumerate(leads, 2):
        for col_idx, (field, _, _) in enumerate(LEAD_COLUMNS, 1):
            value = lead.get(field, "")
            if value is None:
                value = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=str(value))
            cell.border = _THIN_BORDER

    # Freeze header row + auto-filter
    ws.freeze_panes = "A2"
    if leads:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(LEAD_COLUMNS))}{len(leads) + 1}"

    # Summary sheet
    ws_summary = wb.create_sheet("Summary")
    ws_summary["A1"] = "Export Date"
    ws_summary["B1"] = datetime.now().strftime("%d.%m.%Y %H:%M")
    ws_summary["A2"] = "Total Leads"
    ws_summary["B2"] = len(leads)
    ws_summary["A3"] = "Export File"
    ws_summary["B3"] = output.name
    for r in range(1, 4):
        ws_summary.cell(row=r, column=1).font = Font(bold=True)
    ws_summary.column_dimensions["A"].width = 15
    ws_summary.column_dimensions["B"].width = 40

    wb.save(str(output))
    logger.info(f"Exported {len(leads)} leads to Excel: {output}")
    return output
